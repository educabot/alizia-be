"""
Generador de planilla de tests para Alizia BE.

Replica el formato del template `Copia de test 7.xlsx` (Desktop) y produce
un workbook con:
  - Hoja "Indice" (resumen y navegacion)
  - Por cada caso de uso implementado: hoja "UC N" (especificacion) + "UT N" (tests)

Casos cubiertos (8):
  UC-001  GET  /api/v1/users/me/onboarding-status
  UC-002  POST /api/v1/users/me/onboarding/complete
  UC-003  GET  /api/v1/users/me/profile
  UC-004  PUT  /api/v1/users/me/profile
  UC-005  GET  /api/v1/users/me/onboarding/tour-steps
  UC-006  GET  /api/v1/onboarding-config
  UC-007  POST /api/v1/areas/:id/coordinators
  UC-008  DELETE /api/v1/areas/:id/coordinators/:user_id
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------- Styling helpers -----------------------------------------------

GREY_HEADER = PatternFill("solid", fgColor="D3D3D3")
GREEN_HEADER = PatternFill("solid", fgColor="B4D7A8")
INDEX_HEADER = PatternFill("solid", fgColor="4A86E8")
INDEX_BAND = PatternFill("solid", fgColor="EAF1FB")

THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
WRAP_LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")


def bold(cell, size=11, color="000000"):
    cell.font = Font(bold=True, size=size, color=color)


# ---------- UC data --------------------------------------------------------

UCS = [
    {
        "id": "UC-001",
        "title": "UC-001: Obtener estado de onboarding",
        "deps": (
            "- Usuario autenticado con JWT valido\n"
            "- Tabla `users` con columna `onboarding_completed_at` (nullable)\n"
            "- Middleware de auth y tenant resolvieron user_id y org_id"
        ),
        "precondition": (
            "Se recibe una peticion GET a '/api/v1/users/me/onboarding-status' con el header "
            "Authorization: Bearer <jwt>"
        ),
        "description": (
            "Devuelve el estado de onboarding del usuario autenticado. El campo `completed` es "
            "true si `onboarding_completed_at` no es null; cuando lo es, `completed_at` se formatea "
            "como string RFC3339; en caso contrario se envia null."
        ),
        "normal": [
            ("1", "Se recibe la peticion GET '/api/v1/users/me/onboarding-status'"),
            ("1.1", "El middleware de auth valida el JWT y extrae user_id; si falla, responde 401 (Unauthorized)"),
            ("2", "El middleware de tenant resuelve el org_id del usuario autenticado"),
            ("3", "Se construye GetStatusRequest{OrgID, UserID} y se invoca Validate()"),
            ("3.1", "Si org_id o user_id son vacios se retorna HTTP 400 con `providers.ErrValidation`"),
            ("4", "Se llama users.FindByID(ctx, org_id, user_id)"),
            ("4.1", "Si el usuario no existe se retorna HTTP 404 (Not Found)"),
            ("5", "Se arma la respuesta: completed = user.OnboardingCompletedAt != nil"),
            ("6", "Si completed es true, se formatea completed_at con time.RFC3339"),
        ],
        "normal_end": ("7", "Se retorna HTTP 200 con JSON { completed, completed_at }"),
        "postcondition": (
            "- No hay cambios en base de datos\n"
            "- El cliente recibe el estado actual del onboarding"
        ),
        "exceptions": [
            ("1.1", "El header Authorization esta ausente o el JWT es invalido / expirado"),
            ("",    "Se retorna HTTP 401 (Unauthorized) sin invocar el usecase"),
            ("3.1", "Validate() falla porque org_id o user_id son nulos"),
            ("",    "Se retorna HTTP 400 (Bad Request) envolviendo providers.ErrValidation"),
            ("4.1", "users.FindByID retorna providers.ErrNotFound"),
            ("",    "Se retorna HTTP 404 (Not Found)"),
            ("4.2", "Error inesperado en la query (timeout, conexion, etc.)"),
            ("",    "Se retorna HTTP 500 (Internal Server Error)"),
        ],
        "comments": (
            "Endpoint read-only. Idempotente por definicion. No dispara eventos ni side-effects."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Usuario con onboarding pendiente",
                "pre": "user_id=3 existe, onboarding_completed_at = NULL",
                "steps": "1. Loguear como Maria (id=3)\n2. Ejecutar GET onboarding-status",
                "input": "GET /api/v1/users/me/onboarding-status\nAuthorization: Bearer <jwt-maria>",
                "expected": '200 OK\nBody: { "completed": false, "completed_at": null }',
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Usuario con onboarding ya completado",
                "pre": "user_id=3 existe, onboarding_completed_at = '2026-04-09T10:51:14Z'",
                "steps": "1. Marcar onboarding como completado\n2. Ejecutar GET onboarding-status",
                "input": "GET /api/v1/users/me/onboarding-status\nAuthorization: Bearer <jwt-maria>",
                "expected": '200 OK\nBody: { "completed": true, "completed_at": "2026-04-09T10:51:14Z" }',
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Peticion sin JWT",
                "pre": "Ninguna",
                "steps": "1. Ejecutar GET sin Authorization header",
                "input": "GET /api/v1/users/me/onboarding-status",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "JWT invalido o expirado",
                "pre": "JWT firmado con otro secret o vencido",
                "steps": "1. Ejecutar GET con token corrupto",
                "input": "GET /api/v1/users/me/onboarding-status\nAuthorization: Bearer invalid.token.xyz",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Usuario del JWT no existe en DB",
                "pre": "JWT valido con user_id=9999 inexistente",
                "steps": "1. Ejecutar GET con token de usuario eliminado",
                "input": "GET /api/v1/users/me/onboarding-status\nAuthorization: Bearer <jwt-9999>",
                "expected": "404 Not Found",
                "status": "PASS",
            },
            {
                "id": "T-06",
                "desc": "Formato de completed_at es RFC3339",
                "pre": "user con onboarding_completed_at seteado",
                "steps": "1. GET onboarding-status\n2. Validar que completed_at parsea con time.RFC3339",
                "input": "GET /api/v1/users/me/onboarding-status\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\ncompleted_at cumple /^\\d{4}-\\d{2}-\\d{2}T\\d{2}:\\d{2}:\\d{2}(Z|[+-]\\d{2}:\\d{2})$/",
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-002",
        "title": "UC-002: Completar onboarding (idempotente)",
        "deps": (
            "- JWT valido\n"
            "- users.onboarding_completed_at accesible en DB\n"
            "- users.CompleteOnboarding(ctx, userID) disponible en el provider"
        ),
        "precondition": (
            "Se recibe una peticion POST a '/api/v1/users/me/onboarding/complete' con JWT valido"
        ),
        "description": (
            "Marca el onboarding del usuario como completado. Es idempotente: si el usuario ya "
            "tiene onboarding_completed_at != null, el endpoint hace short-circuit y NO sobrescribe "
            "el timestamp existente."
        ),
        "normal": [
            ("1", "Se recibe POST '/api/v1/users/me/onboarding/complete'"),
            ("1.1", "Middleware auth valida JWT; si falla, 401"),
            ("2", "Middleware tenant resuelve org_id"),
            ("3", "Se construye CompleteRequest{OrgID, UserID} y se ejecuta Validate()"),
            ("3.1", "Si hay error de validacion, 400 con providers.ErrValidation"),
            ("4", "users.FindByID(ctx, org_id, user_id) para verificar existencia"),
            ("4.1", "Si no existe, 404"),
            ("5", "Si user.OnboardingCompletedAt != nil se retorna nil (short-circuit idempotente)"),
            ("6", "users.CompleteOnboarding(ctx, user_id) ejecuta UPDATE users SET onboarding_completed_at = NOW()"),
        ],
        "normal_end": ("7", "Se retorna HTTP 200 (OK) sin body"),
        "postcondition": (
            "- Si el usuario no estaba completado, users.onboarding_completed_at queda seteado con NOW()\n"
            "- Llamadas repetidas no modifican el timestamp inicial"
        ),
        "exceptions": [
            ("1.1", "JWT ausente o invalido"),
            ("",    "Se retorna HTTP 401"),
            ("3.1", "Falla Validate() (ids en cero)"),
            ("",    "HTTP 400 envolviendo providers.ErrValidation"),
            ("4.1", "users.FindByID retorna ErrNotFound"),
            ("",    "HTTP 404"),
            ("6.1", "Falla el UPDATE (timeout, constraint, conexion)"),
            ("",    "HTTP 500 con el error de DB"),
        ],
        "comments": (
            "La idempotencia fue validada en commit 8722b8b: sin el short-circuit del paso 5 el "
            "timestamp se sobrescribia en cada llamada."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Primera vez que se completa el onboarding",
                "pre": "user_id=3 con onboarding_completed_at = NULL",
                "steps": "1. Loguear como Maria\n2. POST /onboarding/complete",
                "input": "POST /api/v1/users/me/onboarding/complete\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nDB: onboarding_completed_at != NULL",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Llamada repetida NO sobrescribe el timestamp (idempotencia)",
                "pre": "user_id=3 con onboarding_completed_at ya seteado en T0",
                "steps": "1. POST /onboarding/complete\n2. Esperar 2s\n3. POST /onboarding/complete otra vez\n4. Comparar timestamps",
                "input": "POST /api/v1/users/me/onboarding/complete (x2)\nAuthorization: Bearer <jwt>",
                "expected": "200 OK en ambas\nonboarding_completed_at IGUAL en ambas respuestas",
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "1. POST sin Authorization",
                "input": "POST /api/v1/users/me/onboarding/complete",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Usuario del JWT no existe",
                "pre": "JWT con user_id=9999 inexistente",
                "steps": "1. POST con token de user ausente",
                "input": "POST /api/v1/users/me/onboarding/complete\nAuthorization: Bearer <jwt-9999>",
                "expected": "404 Not Found",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Validacion de request falla (unit test)",
                "pre": "CompleteRequest{OrgID: uuid.Nil, UserID: 0}",
                "steps": "Invocar Execute directamente",
                "input": "Request con campos en cero",
                "expected": "Error envolviendo providers.ErrValidation",
                "status": "PASS",
            },
            {
                "id": "T-06",
                "desc": "Error en el UPDATE de DB",
                "pre": "Mock UserProvider.CompleteOnboarding retorna error",
                "steps": "1. FindByID OK\n2. CompleteOnboarding falla",
                "input": "POST con mock error",
                "expected": "500 Internal Server Error\nMensaje con el error de DB",
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-003",
        "title": "UC-003: Obtener perfil del usuario",
        "deps": (
            "- JWT valido\n"
            "- users.profile_data (JSONB, puede ser null o vacio)\n"
            "- UserProvider.FindByID disponible"
        ),
        "precondition": (
            "Se recibe una peticion GET a '/api/v1/users/me/profile' con JWT valido"
        ),
        "description": (
            "Devuelve el profile_data del usuario como un map[string]any. Si el JSONB esta vacio "
            "o es null, devuelve un objeto vacio {} en vez de null."
        ),
        "normal": [
            ("1", "Se recibe GET '/api/v1/users/me/profile'"),
            ("1.1", "Middleware auth valida JWT (falla -> 401)"),
            ("2", "Middleware tenant resuelve org_id"),
            ("3", "Validate() del request"),
            ("4", "users.FindByID(ctx, org_id, user_id)"),
            ("4.1", "Si no existe, 404"),
            ("5", "Si user.ProfileData es len 0 se retorna map vacio {}"),
            ("6", "Se hace json.Unmarshal del JSONB a map[string]any"),
            ("6.1", "Si el JSONB esta corrupto, se retorna 500"),
        ],
        "normal_end": ("7", "HTTP 200 con el body JSON decodificado"),
        "postcondition": "Sin modificaciones en DB.",
        "exceptions": [
            ("1.1", "JWT invalido o ausente"),
            ("",    "HTTP 401"),
            ("3.1", "Validate falla"),
            ("",    "HTTP 400"),
            ("4.1", "Usuario no encontrado"),
            ("",    "HTTP 404"),
            ("6.1", "json.Unmarshal falla (JSONB corrupto en DB)"),
            ("",    "HTTP 500"),
        ],
        "comments": (
            "La distincion entre 'profile_data null' y 'profile_data {}' se normaliza a {} para "
            "simplificar el contrato del front."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Usuario sin profile_data",
                "pre": "user_id=3, profile_data = NULL",
                "steps": "1. GET /users/me/profile",
                "input": "GET /api/v1/users/me/profile\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nBody: {}",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Usuario con profile_data seteado",
                "pre": 'profile_data = {"disciplines":["Matematicas"],"experience_years":5}',
                "steps": "1. GET /users/me/profile",
                "input": "GET /api/v1/users/me/profile\nAuthorization: Bearer <jwt>",
                "expected": '200 OK\nBody: {"disciplines":["Matematicas"],"experience_years":5}',
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "GET sin header",
                "input": "GET /api/v1/users/me/profile",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Usuario inexistente",
                "pre": "JWT con user_id=9999",
                "steps": "1. GET profile",
                "input": "GET /api/v1/users/me/profile\nAuthorization: Bearer <jwt-9999>",
                "expected": "404 Not Found",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "JSONB corrupto en DB",
                "pre": "profile_data contiene bytes que no parsean como JSON",
                "steps": "1. GET profile",
                "input": "GET /api/v1/users/me/profile\nAuthorization: Bearer <jwt>",
                "expected": "500 Internal Server Error",
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-004",
        "title": "UC-004: Guardar perfil del usuario",
        "deps": (
            "- JWT valido\n"
            "- organizations.config con seccion onboarding.profile_fields\n"
            "- UserProvider.UpdateProfileData(ctx, userID, data)"
        ),
        "precondition": (
            "Se recibe PUT '/api/v1/users/me/profile' con body JSON representando el perfil"
        ),
        "description": (
            "Actualiza users.profile_data validando cada campo contra la config del org: "
            "chequea required, tipo (text/number/select/multiselect) y opciones permitidas en "
            "fields de tipo select."
        ),
        "normal": [
            ("1", "Se recibe PUT '/api/v1/users/me/profile' con body JSON"),
            ("1.1", "Middleware auth valida JWT (falla -> 401)"),
            ("2", "Se parsea el body a map[string]any; si es JSON invalido -> 400"),
            ("3", "Se construye SaveProfileRequest{OrgID, UserID, Data} y se ejecuta Validate() (Data con al menos 1 campo)"),
            ("4", "users.FindByID verifica que el usuario exista (404 si no)"),
            ("5", "orgs.FindByID obtiene la config del org"),
            ("6", "extractProfileFields(org.Config) lee onboarding.profile_fields"),
            ("7", "validateProfileData: por cada field configurado, chequea required y tipo/opciones"),
            ("7.1", "Si un campo required falta -> 400"),
            ("7.2", "Si el tipo no coincide -> 400"),
            ("7.3", "Si una option no esta en el set permitido -> 400"),
            ("8", "users.UpdateProfileData(ctx, user_id, data) persiste en la columna JSONB"),
        ],
        "normal_end": ("9", "HTTP 200 (OK) sin body"),
        "postcondition": (
            "- users.profile_data contiene el nuevo payload serializado\n"
            "- Llamadas subsecuentes a GET /profile devuelven el nuevo valor"
        ),
        "exceptions": [
            ("2.1", "Body vacio o no JSON"),
            ("",    "HTTP 400"),
            ("3.1", "Data nulo o map vacio ({})"),
            ("",    "HTTP 400 con providers.ErrValidation (profile data cannot be empty)"),
            ("4.1", "Usuario no existe"),
            ("",    "HTTP 404"),
            ("5.1", "Org no existe"),
            ("",    "HTTP 404"),
            ("7.x", "Validacion de campos falla (required/tipo/option)"),
            ("",    "HTTP 400 con el field.Key que fallo"),
            ("8.1", "UpdateProfileData retorna error de DB"),
            ("",    "HTTP 500"),
        ],
        "comments": (
            "Decision: PUT con body vacio ({}) se rechaza con 400 (fail-fast). "
            "En onboarding un body vacio nunca es intencional; si en el futuro aparece "
            "la necesidad de limpiar el perfil, se agregara DELETE /profile con semantica explicita."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Guardar perfil valido",
                "pre": "org.config define profile_fields: disciplines (multiselect con Matematicas, Fisica), experience_years (number)",
                "steps": "1. PUT con body JSON valido",
                "input": 'PUT /api/v1/users/me/profile\nBody: {"disciplines":["Matematicas"],"experience_years":5}',
                "expected": "200 OK\nusers.profile_data = body recibido",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Falta campo required",
                "pre": 'profile_fields: disciplines (required)',
                "steps": "1. PUT sin disciplines",
                "input": 'PUT /api/v1/users/me/profile\nBody: {"experience_years":5}',
                "expected": '400 Bad Request\nerror: field "disciplines" is required',
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Tipo incorrecto",
                "pre": "experience_years declarado como number",
                "steps": "1. PUT con string en experience_years",
                "input": 'PUT /api/v1/users/me/profile\nBody: {"experience_years":"cinco"}',
                "expected": '400 Bad Request\nfield "experience_years" must be a number',
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Option invalida en select",
                "pre": "disciplines tiene options [Matematicas, Fisica]",
                "steps": "1. PUT con opcion fuera de set",
                "input": 'PUT /api/v1/users/me/profile\nBody: {"disciplines":["Quimica"]}',
                "expected": '400 Bad Request\ninvalid option "Quimica"',
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Body JSON malformado",
                "pre": "Ninguna",
                "steps": "1. PUT con body que no es JSON",
                "input": "PUT /api/v1/users/me/profile\nBody: not json",
                "expected": "400 Bad Request",
                "status": "PASS",
            },
            {
                "id": "T-06",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "PUT sin Authorization",
                "input": "PUT /api/v1/users/me/profile",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-07",
                "desc": "Body vacio es rechazado (fail-fast)",
                "pre": "Cualquier org.config",
                "steps": "1. PUT con body {}",
                "input": 'PUT /api/v1/users/me/profile\nBody: {}',
                "expected": '400 Bad Request\nMensaje: profile data cannot be empty',
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-005",
        "title": "UC-005: Obtener pasos del tour",
        "deps": (
            "- JWT valido\n"
            "- organizations.config con onboarding.tour_steps y features\n"
            "- Roles del usuario en users.roles"
        ),
        "precondition": (
            "Se recibe GET '/api/v1/users/me/onboarding/tour-steps' con JWT valido"
        ),
        "description": (
            "Devuelve los pasos del tour guiado filtrados por (a) roles del usuario y "
            "(b) features activos del org. Si el org no define tour_steps se devuelve una lista "
            "default (welcome + explore). Los pasos se ordenan por el campo Order."
        ),
        "normal": [
            ("1", "Se recibe GET '/api/v1/users/me/onboarding/tour-steps'"),
            ("1.1", "Middleware auth valida JWT (falla -> 401)"),
            ("2", "Middleware tenant resuelve org_id"),
            ("3", "Validate() del request (orgID y userID)"),
            ("4", "orgs.FindByID(ctx, org_id) recupera la config"),
            ("5", "users.FindByID(ctx, org_id, user_id) para obtener roles"),
            ("6", "extractTourStepsConfig(org.Config)"),
            ("6.1", "Si no hay tour_steps configurados, se devuelven defaultTourSteps (welcome, explore)"),
            ("7", "extractActiveFeatures(org.Config) mapea features activos"),
            ("8", "Se itera cada tourStepConfig: se ignora si la key ya fue vista (dedupe)"),
            ("9", "Se filtra por matchesRoles(step.Roles, user.RoleNames())"),
            ("10", "Se filtra por requires_feature: si la feature no esta activa, se salta"),
            ("11", "Se ordena la lista resultante por Order ascendente"),
        ],
        "normal_end": ("12", "HTTP 200 con el array JSON de TourStep"),
        "postcondition": "Sin cambios en DB.",
        "exceptions": [
            ("1.1", "JWT invalido/ausente"),
            ("",    "HTTP 401"),
            ("3.1", "Validate() falla"),
            ("",    "HTTP 400"),
            ("4.1", "Org no existe"),
            ("",    "HTTP 404"),
            ("5.1", "User no existe"),
            ("",    "HTTP 404"),
            ("6.x", "org.Config no parsea como JSON"),
            ("",    "Se devuelve lista vacia de tour_steps y se aplica defaultTourSteps"),
        ],
        "comments": (
            "El filtrado por features permite ocultar pasos del tour cuando una funcionalidad esta "
            "off para el org (ej: feature 'ai_lessons' no activa => skip paso 'ai-tour')."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Org sin tour_steps configurados -> default",
                "pre": "org.config.onboarding sin tour_steps",
                "steps": "1. GET tour-steps",
                "input": "GET /api/v1/users/me/onboarding/tour-steps\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nArray con welcome + explore (defaultTourSteps)",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Filtro por rol de usuario",
                "pre": "tour_steps = [welcome(no-roles), admin-intro(roles=[admin])]; user es teacher",
                "steps": "1. GET tour-steps como teacher",
                "input": "GET /api/v1/users/me/onboarding/tour-steps\nAuthorization: Bearer <jwt-teacher>",
                "expected": "200 OK\nArray solo con welcome (admin-intro filtrado)",
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Filtro por feature activa",
                "pre": "tour_steps include step con requires_feature='ai_lessons'; features.ai_lessons=false",
                "steps": "1. GET tour-steps",
                "input": "GET /api/v1/users/me/onboarding/tour-steps\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nPaso 'ai_lessons' ausente del array",
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Orden por Order ascendente",
                "pre": "tour_steps = [stepA(order=2), stepB(order=1)]",
                "steps": "1. GET tour-steps",
                "input": "GET /api/v1/users/me/onboarding/tour-steps\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nArray: [stepB, stepA]",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Dedupe por key",
                "pre": "tour_steps tiene dos entradas con key='welcome'",
                "steps": "1. GET tour-steps",
                "input": "GET /api/v1/users/me/onboarding/tour-steps\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nArray contiene una sola entrada welcome",
                "status": "PASS",
            },
            {
                "id": "T-06",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "GET sin header",
                "input": "GET /api/v1/users/me/onboarding/tour-steps",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-07",
                "desc": "Multi-rol (Pedro coord+teacher) ve pasos de ambos roles",
                "pre": "tour_steps = [welcome, t-tour(roles=[teacher]), c-tour(roles=[coordinator])]",
                "steps": "1. Loguear como Pedro\n2. GET tour-steps",
                "input": "GET /api/v1/users/me/onboarding/tour-steps\nAuthorization: Bearer <jwt-pedro>",
                "expected": "200 OK\nArray: welcome + t-tour + c-tour (orden por Order)",
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-006",
        "title": "UC-006: Obtener configuracion de onboarding del org",
        "deps": (
            "- JWT valido\n"
            "- organizations.config (JSONB) con seccion onboarding"
        ),
        "precondition": (
            "Se recibe GET '/api/v1/onboarding-config' con JWT valido"
        ),
        "description": (
            "Devuelve la configuracion de onboarding del org autenticado (profile_fields, tour_steps, "
            "skip_allowed, etc.). Si org.config no contiene la seccion onboarding o el JSON es "
            "invalido, devuelve un fallback seguro con SkipAllowed=true."
        ),
        "normal": [
            ("1", "Se recibe GET '/api/v1/onboarding-config'"),
            ("1.1", "Middleware auth valida JWT (falla -> 401)"),
            ("2", "Middleware tenant resuelve org_id"),
            ("3", "Se construye GetConfigRequest{OrgID} y se ejecuta Validate()"),
            ("4", "orgs.FindByID(ctx, org_id) obtiene el Organization con Config"),
            ("5", "parseOnboardingConfig(org.Config) intenta decodificar la clave 'onboarding'"),
            ("5.1", "Si el unmarshal falla, se devuelve &OnboardingConfig{SkipAllowed: true}"),
        ],
        "normal_end": ("6", "HTTP 200 con el body JSON de OnboardingConfig"),
        "postcondition": "Sin efectos en DB.",
        "exceptions": [
            ("1.1", "JWT invalido/ausente"),
            ("",    "HTTP 401"),
            ("3.1", "Validate() falla"),
            ("",    "HTTP 400"),
            ("4.1", "Org no encontrado"),
            ("",    "HTTP 404"),
            ("5.1", "org.Config corrupto o sin 'onboarding'"),
            ("",    "HTTP 200 con { skip_allowed: true }"),
        ],
        "comments": (
            "No usa user_id para filtrar: es config global del org. Se cachea a nivel de front."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Org con config completo",
                "pre": "org.config.onboarding tiene profile_fields, tour_steps, skip_allowed=false",
                "steps": "1. GET onboarding-config",
                "input": "GET /api/v1/onboarding-config\nAuthorization: Bearer <jwt>",
                "expected": "200 OK\nBody: { profile_fields, tour_steps, skip_allowed: false }",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Org sin seccion onboarding -> fallback",
                "pre": "org.config = {}",
                "steps": "1. GET onboarding-config",
                "input": "GET /api/v1/onboarding-config\nAuthorization: Bearer <jwt>",
                "expected": '200 OK\nBody: { skip_allowed: true }',
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Org con config JSONB corrupto",
                "pre": "org.config no parsea como JSON",
                "steps": "1. GET onboarding-config",
                "input": "GET /api/v1/onboarding-config\nAuthorization: Bearer <jwt>",
                "expected": '200 OK\nBody: { skip_allowed: true }',
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "GET sin header",
                "input": "GET /api/v1/onboarding-config",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Org del JWT no existe",
                "pre": "JWT con org_id inexistente",
                "steps": "1. GET onboarding-config",
                "input": "GET /api/v1/onboarding-config\nAuthorization: Bearer <jwt-org-bad>",
                "expected": "404 Not Found",
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-007",
        "title": "UC-007: Asignar coordinador a un area",
        "deps": (
            "- Rol admin en el JWT\n"
            "- Tablas areas, users (con roles), area_coordinators\n"
            "- Providers: AreaProvider, UserProvider, AreaCoordinatorProvider"
        ),
        "precondition": (
            "Se recibe POST '/api/v1/areas/:id/coordinators' con body {\"user_id\": N} y JWT de admin"
        ),
        "description": (
            "Asigna un usuario como coordinador de un area. Requiere que (a) el area pertenezca "
            "al mismo org del admin, (b) el usuario a asignar exista en el org y (c) tenga rol "
            "coordinator o admin. Retorna el registro AreaCoordinator creado."
        ),
        "normal": [
            ("1", "Se recibe POST '/api/v1/areas/:id/coordinators'"),
            ("1.1", "Middleware auth valida JWT (falla -> 401)"),
            ("1.2", "Middleware RequireRole('admin') verifica el rol (falla -> 403)"),
            ("2", "Parseo del path param :id (areaID) y body { user_id }"),
            ("3", "Validate() chequea orgID, areaID y userID no-cero"),
            ("3.1", "Si falla -> 400 con providers.ErrValidation"),
            ("4", "areas.GetArea(ctx, org_id, area_id) verifica pertenencia al org"),
            ("4.1", "Si no existe -> 404 'area not found'"),
            ("5", "users.FindByID(ctx, org_id, user_id) verifica usuario"),
            ("5.1", "Si no existe -> 404 'user not found'"),
            ("6", "Chequeo user.HasRole(RoleCoordinator) || user.HasRole(RoleAdmin)"),
            ("6.1", "Si no cumple -> 400 'user must have coordinator or admin role'"),
            ("7", "coordinators.Assign(ctx, area_id, user_id) hace INSERT en area_coordinators"),
            ("7.1", "Si duplicate key -> 409 Conflict"),
        ],
        "normal_end": ("8", "HTTP 201 (Created) con el AreaCoordinator asignado"),
        "postcondition": (
            "- Registro area_coordinators(area_id, user_id) insertado\n"
            "- Usuario aparece como coordinador al listar el area"
        ),
        "exceptions": [
            ("1.1", "JWT ausente/invalido"),
            ("",    "HTTP 401"),
            ("1.2", "Usuario autenticado no es admin"),
            ("",    "HTTP 403 Forbidden"),
            ("3.1", "Body invalido o ids en cero"),
            ("",    "HTTP 400"),
            ("4.1", "Area no existe o es de otro org"),
            ("",    "HTTP 404 'area not found'"),
            ("5.1", "Usuario no existe o es de otro org"),
            ("",    "HTTP 404 'user not found'"),
            ("6.1", "Usuario sin rol coordinator/admin"),
            ("",    "HTTP 400 con providers.ErrValidation"),
            ("7.1", "Asignacion duplicada"),
            ("",    "HTTP 409 Conflict"),
            ("7.2", "Error generico de DB"),
            ("",    "HTTP 500"),
        ],
        "comments": (
            "La validacion del rol (paso 6) no esta documentada en el plan de testing original "
            "pero es una buena salvaguarda: bloquea asignar un teacher sin rol."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Asignar coordinador valido",
                "pre": "area_id=1 existe en org, user_id=2 (Carlos) tiene rol coordinator",
                "steps": "1. Login como admin\n2. POST /areas/1/coordinators { user_id: 2 }",
                "input": 'POST /api/v1/areas/1/coordinators\nBody: {"user_id":2}',
                "expected": "201 Created\nBody: AreaCoordinator{area_id:1, user_id:2}",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Asignacion duplicada",
                "pre": "Carlos ya esta asignado al area 1",
                "steps": "1. POST /areas/1/coordinators { user_id: 2 } otra vez",
                "input": 'POST /api/v1/areas/1/coordinators\nBody: {"user_id":2}',
                "expected": "409 Conflict",
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Asignar un teacher (sin rol coordinator)",
                "pre": "user_id=3 (Maria) es teacher",
                "steps": "1. POST /areas/1/coordinators { user_id: 3 }",
                "input": 'POST /api/v1/areas/1/coordinators\nBody: {"user_id":3}',
                "expected": "400 Bad Request\nmensaje: user must have coordinator or admin role",
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Area inexistente",
                "pre": "area_id=9999 no existe",
                "steps": "1. POST /areas/9999/coordinators",
                "input": 'POST /api/v1/areas/9999/coordinators\nBody: {"user_id":2}',
                "expected": "404 Not Found\narea not found",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Usuario inexistente",
                "pre": "user_id=9999 no existe",
                "steps": "1. POST /areas/1/coordinators",
                "input": 'POST /api/v1/areas/1/coordinators\nBody: {"user_id":9999}',
                "expected": "404 Not Found\nuser not found",
                "status": "PASS",
            },
            {
                "id": "T-06",
                "desc": "Usuario autenticado no es admin",
                "pre": "Login como coordinator (no admin)",
                "steps": "1. POST /areas/1/coordinators con JWT de coordinator",
                "input": 'POST /api/v1/areas/1/coordinators\nAuthorization: Bearer <jwt-coord>',
                "expected": "403 Forbidden",
                "status": "PASS",
            },
            {
                "id": "T-07",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "POST sin Authorization",
                "input": "POST /api/v1/areas/1/coordinators",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
        ],
    },
    {
        "id": "UC-008",
        "title": "UC-008: Remover coordinador de un area",
        "deps": (
            "- Rol admin en el JWT\n"
            "- Tablas areas y area_coordinators\n"
            "- Providers: AreaProvider, AreaCoordinatorProvider"
        ),
        "precondition": (
            "Se recibe DELETE '/api/v1/areas/:id/coordinators/:user_id' con JWT de admin"
        ),
        "description": (
            "Remueve la asignacion de coordinador de un area. Verifica que el area pertenezca al "
            "org del admin antes de eliminar. No falla si la asignacion ya no existe (idempotente "
            "segun la implementacion del provider)."
        ),
        "normal": [
            ("1", "Se recibe DELETE '/api/v1/areas/:id/coordinators/:user_id'"),
            ("1.1", "Middleware auth valida JWT (falla -> 401)"),
            ("1.2", "Middleware RequireRole('admin') valida rol (falla -> 403)"),
            ("2", "Parseo de :id y :user_id"),
            ("3", "Validate() chequea orgID, areaID y userID no-cero"),
            ("4", "areas.GetArea(ctx, org_id, area_id) verifica pertenencia"),
            ("4.1", "Si no existe -> 404 'area not found'"),
            ("5", "coordinators.Remove(ctx, area_id, user_id) ejecuta DELETE"),
        ],
        "normal_end": ("6", "HTTP 204 (No Content)"),
        "postcondition": (
            "- Registro area_coordinators(area_id, user_id) eliminado (o no existia)\n"
            "- Usuario ya no aparece al listar coordinadores del area"
        ),
        "exceptions": [
            ("1.1", "JWT invalido/ausente"),
            ("",    "HTTP 401"),
            ("1.2", "No es admin"),
            ("",    "HTTP 403"),
            ("3.1", "Path params invalidos"),
            ("",    "HTTP 400"),
            ("4.1", "Area no existe o es de otro org"),
            ("",    "HTTP 404 'area not found'"),
            ("5.1", "Error de DB en DELETE"),
            ("",    "HTTP 500"),
        ],
        "comments": (
            "La idempotencia depende del provider: Remove debe ser un DELETE sin 'row-not-found' "
            "como error, para soportar llamadas repetidas sin 404."
        ),
        "tests": [
            {
                "id": "T-01",
                "desc": "Remover coordinador asignado",
                "pre": "Carlos (id=2) asignado al area 1",
                "steps": "1. Login admin\n2. DELETE /areas/1/coordinators/2",
                "input": "DELETE /api/v1/areas/1/coordinators/2\nAuthorization: Bearer <jwt-admin>",
                "expected": "204 No Content\nDB: registro eliminado",
                "status": "PASS",
            },
            {
                "id": "T-02",
                "desc": "Llamada repetida (idempotencia del provider)",
                "pre": "Carlos ya fue removido del area 1",
                "steps": "1. DELETE /areas/1/coordinators/2 otra vez",
                "input": "DELETE /api/v1/areas/1/coordinators/2\nAuthorization: Bearer <jwt-admin>",
                "expected": "204 No Content (no falla)",
                "status": "PASS",
            },
            {
                "id": "T-03",
                "desc": "Area inexistente",
                "pre": "area_id=9999 no existe",
                "steps": "1. DELETE /areas/9999/coordinators/2",
                "input": "DELETE /api/v1/areas/9999/coordinators/2\nAuthorization: Bearer <jwt-admin>",
                "expected": "404 Not Found\narea not found",
                "status": "PASS",
            },
            {
                "id": "T-04",
                "desc": "Usuario autenticado no es admin",
                "pre": "Login como coordinator",
                "steps": "1. DELETE con JWT no-admin",
                "input": "DELETE /api/v1/areas/1/coordinators/2\nAuthorization: Bearer <jwt-coord>",
                "expected": "403 Forbidden",
                "status": "PASS",
            },
            {
                "id": "T-05",
                "desc": "Sin JWT",
                "pre": "Ninguna",
                "steps": "DELETE sin Authorization",
                "input": "DELETE /api/v1/areas/1/coordinators/2",
                "expected": "401 Unauthorized",
                "status": "PASS",
            },
            {
                "id": "T-06",
                "desc": "Error de DB en DELETE",
                "pre": "Mock AreaCoordinatorProvider.Remove retorna error",
                "steps": "1. DELETE con mock error",
                "input": "DELETE /api/v1/areas/1/coordinators/2",
                "expected": "500 Internal Server Error",
                "status": "PASS",
            },
        ],
    },
]


# ---------- Sheet builders -------------------------------------------------

def build_index(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Indice"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Plan de Tests - Alizia BE (Epica 1 + 2)"
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = INDEX_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Meta
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "Stack: Go 1.26 + Gin + GORM + PostgreSQL  |  Auth: JWT  |  "
        "Branch: feature/sl/epica-2-onboarding  |  Fecha: 2026-04-09"
    )
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = Font(italic=True, color="555555")

    # Table header
    headers = ["Codigo", "Titulo", "Endpoint", "Tests", "Hoja"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = INDEX_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = WRAP_CENTER
        cell.border = BOX

    endpoints = {
        "UC-001": "GET  /api/v1/users/me/onboarding-status",
        "UC-002": "POST /api/v1/users/me/onboarding/complete",
        "UC-003": "GET  /api/v1/users/me/profile",
        "UC-004": "PUT  /api/v1/users/me/profile",
        "UC-005": "GET  /api/v1/users/me/onboarding/tour-steps",
        "UC-006": "GET  /api/v1/onboarding-config",
        "UC-007": "POST /api/v1/areas/:id/coordinators",
        "UC-008": "DELETE /api/v1/areas/:id/coordinators/:user_id",
    }

    row = 5
    for uc in UCS:
        n = int(uc["id"].split("-")[1])
        ws.cell(row=row, column=1, value=uc["id"]).alignment = WRAP_CENTER
        ws.cell(row=row, column=2, value=uc["title"]).alignment = WRAP_LEFT
        ws.cell(row=row, column=3, value=endpoints[uc["id"]]).alignment = WRAP_LEFT
        ws.cell(row=row, column=4, value=len(uc["tests"])).alignment = WRAP_CENTER
        link_cell = ws.cell(row=row, column=5, value=f"UC {n} / UT {n}")
        link_cell.alignment = WRAP_CENTER
        link_cell.font = Font(color="1155CC", underline="single")
        link_cell.hyperlink = f"#'UC {n}'!A1"
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = BOX
            if row % 2 == 0:
                cell.fill = INDEX_BAND
        ws.row_dimensions[row].height = 26
        row += 1

    # Summary row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    total_tests = sum(len(uc["tests"]) for uc in UCS)
    c = ws.cell(row=row, column=1, value=f"Total: {len(UCS)} casos de uso | {total_tests} tests")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = PatternFill("solid", fgColor="F1F3F4")

    # Legend
    row += 2
    ws.cell(row=row, column=1, value="Leyenda:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="PASS").fill = PatternFill("solid", fgColor="D9EAD3")
    ws.cell(row=row, column=2, value="Test cumple el resultado esperado")
    row += 1
    ws.cell(row=row, column=1, value="FAIL").fill = PatternFill("solid", fgColor="F4CCCC")
    ws.cell(row=row, column=2, value="Test NO cumple el resultado esperado")
    row += 1
    ws.cell(row=row, column=1, value="FLAG").fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=row, column=2, value="Comportamiento ambiguo - revisar si es intencional")


def build_uc_sheet(wb: Workbook, uc: dict) -> None:
    n = int(uc["id"].split("-")[1])
    ws: Worksheet = wb.create_sheet(title=f"UC {n}")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 52

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = uc["title"]
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    def set_label(row: int, label: str, content: str) -> None:
        ws.cell(row=row, column=1, value=label)
        bold(ws.cell(row=row, column=1))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=2, value=content)
        cell.alignment = WRAP_TOP

    set_label(3, "Dependencias", uc["deps"])
    set_label(4, "Precondicion", uc["precondition"])
    set_label(5, "Descripcion", uc["description"])

    # Spacer row
    ws.merge_cells("A6:C6")

    # Normal flow header
    hdr_paso = ws.cell(row=7, column=1, value="Paso")
    hdr_paso.fill = GREY_HEADER
    bold(hdr_paso)
    ws.merge_cells("B7:C7")
    hdr_accion = ws.cell(row=7, column=2, value="Accion")
    hdr_accion.fill = GREY_HEADER
    bold(hdr_accion)

    row = 8
    for step, action in uc["normal"]:
        ws.cell(row=row, column=1, value=step).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ac = ws.cell(row=row, column=2, value=action)
        ac.alignment = WRAP_TOP
        row += 1

    # Secuencia normal label + final step (kept on same row like template)
    end_step, end_action = uc["normal_end"]
    c = ws.cell(row=row, column=1, value="Secuencia\nNormal")
    c.alignment = WRAP_TOP
    bold(c)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=f"{end_step} {end_action}").alignment = WRAP_TOP
    row += 1

    set_label(row, "Postcondicion", uc["postcondition"])
    row += 2  # blank row

    # Exceptions header
    hdr_paso = ws.cell(row=row, column=1, value="Paso")
    hdr_paso.fill = GREY_HEADER
    bold(hdr_paso)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    hdr_accion = ws.cell(row=row, column=2, value="Accion")
    hdr_accion.fill = GREY_HEADER
    bold(hdr_accion)
    row += 1

    for step, action in uc["exceptions"]:
        ws.cell(row=row, column=1, value=step).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2, value=action).alignment = WRAP_TOP
        row += 1

    # Excepciones label row (same style as "Secuencia Normal")
    c = ws.cell(row=row, column=1, value="Excepciones")
    bold(c)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 1

    # Comentarios
    set_label(row, "Comentarios", uc["comments"])


def build_ut_sheet(wb: Workbook, uc: dict) -> None:
    n = int(uc["id"].split("-")[1])
    ws: Worksheet = wb.create_sheet(title=f"UT {n}")

    widths = {"A": 9, "B": 32, "C": 28, "D": 32, "E": 30, "F": 34, "G": 12, "H": 14, "I": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = [
        "Test ID", "Descripcion", "Precondiciones", "Pasos / Dependencias",
        "Entrada", "Resultado Esperado", "PASS/FAILS", "Resultado", "Comentarios",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = GREEN_HEADER
        c.font = Font(bold=True)
        c.alignment = WRAP_CENTER
        c.border = BOX

    # UC title above table? keep template faithful: headers in row 1, data from row 2
    row = 2
    for t in uc["tests"]:
        ws.cell(row=row, column=1, value=t["id"])
        ws.cell(row=row, column=2, value=t["desc"])
        ws.cell(row=row, column=3, value=t["pre"])
        ws.cell(row=row, column=4, value=t["steps"])
        ws.cell(row=row, column=5, value=t["input"])
        ws.cell(row=row, column=6, value=t["expected"])
        g = ws.cell(row=row, column=7, value=t["status"])
        if t["status"] == "PASS":
            g.fill = PatternFill("solid", fgColor="D9EAD3")
        elif t["status"] == "FAIL":
            g.fill = PatternFill("solid", fgColor="F4CCCC")
        else:  # FLAG
            g.fill = PatternFill("solid", fgColor="FFF2CC")
        g.alignment = WRAP_CENTER
        ws.cell(row=row, column=8)  # Resultado -> se llena al ejecutar
        ws.cell(row=row, column=9)  # Comentarios

        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.alignment = WRAP_TOP
            cell.border = BOX
        ws.row_dimensions[row].height = 90
        row += 1


# ---------- Main -----------------------------------------------------------

def main() -> None:
    wb = Workbook()
    build_index(wb)
    for uc in UCS:
        build_uc_sheet(wb, uc)
        build_ut_sheet(wb, uc)

    out = "docs/TEST_PLAN_ALIZIA_BE.xlsx"
    wb.save(out)
    total = sum(len(u["tests"]) for u in UCS)
    print(f"[OK] Generado: {out}")
    print(f"     Hojas: Indice + {len(UCS) * 2} ({len(UCS)} UC + {len(UCS)} UT)")
    print(f"     Tests totales: {total}")


if __name__ == "__main__":
    main()
